
import streamlit as st
import pandas as pd
import openpyxl
import sqlite3
import os

def show_stock_ohlc_update_page():
    st.title("📈 Stock OHLC Database Update")

    # --- Integrated template download ---
    template_path = "data/Integrated_BTH_Template.xlsx"
    if os.path.exists(template_path):
        with open(template_path, "rb") as f:
            st.sidebar.download_button(
                label="📥 Download BTH Template Sample",
                data=f,
                file_name="BTH_Template_Sample.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.sidebar.info("ℹ️ BTH Template sample file not found.")

    st.sidebar.header("Mode Selection")
    mode = st.sidebar.selectbox("Select Mode", ["Update / Create Stock Database", "Read an Existing Database"])
    db_path = st.sidebar.text_input("SQLite DB Path","ohlc_bbdata.db")

    def parse_excel(file):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        datasets = []
        col_index = 1
        while True:
            stock_cell = ws.cell(row=4, column=col_index)
            if stock_cell.value is None:
                break
            stock_raw = stock_cell.value
            stock = stock_raw.split()[0] if isinstance(stock_raw, str) else stock_raw
            data = []
            row_index = 6
            while True:
                row = [ws.cell(row=row_index, column=col_index + i).value for i in range(7)]
                if all(cell is None for cell in row):
                    break
                data.append(row)
                row_index += 1
            if data:
                df = pd.DataFrame(data, columns=["Date", "Open", "High", "Low", "Close", "Volume", "Value"])
                df.insert(0, "Stock", stock)
                df["Date"] = pd.to_datetime(df["Date"]).dt.date
                datasets.append(df)
            col_index += 8
        return pd.concat(datasets, ignore_index=True) if datasets else pd.DataFrame()

    def save_to_db(df, db_path):
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stock_data (
                Stock TEXT,
                Date TEXT,
                Open REAL,
                High REAL,
                Low REAL,
                Close REAL,
                Volume INTEGER,
                Value REAL,
                VWAP REAL,
                PRIMARY KEY (Stock, Date)
            )
        """)
        existing_keys = pd.read_sql("SELECT Stock, Date FROM stock_data", conn)
        existing_keys["Date"] = pd.to_datetime(existing_keys["Date"]).dt.date
        df_merged = df.merge(existing_keys, on=["Stock", "Date"], how="left", indicator=True)
        new_data = df_merged[df_merged["_merge"] == "left_only"].drop(columns=["_merge"])
        if not new_data.empty:
            new_data[["Stock", "Date", "Open", "High", "Low", "Close", "Volume", "Value", "VWAP"]].to_sql(
                "stock_data", conn, if_exists="append", index=False
            )
        conn.close()
        return new_data

    def read_database(db_path):
        if not os.path.exists(db_path):
            st.error(f"Database file `{db_path}` not found.")
            return pd.DataFrame()
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT * FROM stock_data", conn, parse_dates=["Date"])
        df["Date"] = pd.to_datetime(df["Date"]).dt.date
        conn.close()
        return df

    if mode == "Update / Create Stock Database":
        uploaded_file = st.sidebar.file_uploader("Upload Excel File", type=["xlsx"])
        create_db_btn = st.sidebar.button("Save to Database")

        if uploaded_file:
            parsed_df = parse_excel(uploaded_file)
            if not parsed_df.empty:
                parsed_df["VWAP"] = (parsed_df["Value"] / parsed_df["Volume"]).round(4)
                parsed_df["Volume"] = parsed_df["Volume"].astype("Int64")
                parsed_df["Value"] = parsed_df["Value"].astype("Int64")
                preview_df = parsed_df.copy()
                preview_df["Volume"] = preview_df["Volume"].map("{:,}".format)
                preview_df["Value"] = preview_df["Value"].map("{:,}".format)
                st.subheader("📋 Preview of Uploaded Data with VWAP")
                st.dataframe(preview_df.head(10))

                if create_db_btn:
                    new_rows = save_to_db(parsed_df, db_path)
                    if not new_rows.empty:
                        summary = new_rows.groupby("Stock")["Date"].agg(["min", "max"]).reset_index()
                        summary.columns = ["Stock", "Date From", "Date To"]
                        st.success(f"✅ {len(new_rows)} new records saved to `{db_path}`.")
                        st.subheader("📈 Summary of Newly Added Data")
                        st.dataframe(summary)
                    else:
                        st.info("ℹ️ No new records inserted. All data already exists in the database.")

    elif mode == "Read an Existing Database":
        if "db_loaded" not in st.session_state:
            st.session_state.db_loaded = False
        if "db_data" not in st.session_state:
            st.session_state.db_data = pd.DataFrame()

        read_db_btn = st.sidebar.button("Read Database")
        if read_db_btn:
            df = read_database(db_path)
            if df.empty:
                st.warning("⚠️ No data found in the database.")
                st.session_state.db_loaded = False
            else:
                st.session_state.db_data = df
                st.session_state.db_loaded = True

        if st.session_state.db_loaded:
            df = st.session_state.db_data

            st.sidebar.markdown("---")
            st.sidebar.subheader("📆 Filter Options")

            min_date = df["Date"].min()
            max_date = df["Date"].max()
            date_range = st.sidebar.date_input("Select Date Range", [min_date, max_date], min_value=min_date, max_value=max_date)

            stocks = df["Stock"].unique().tolist()
            use_input_field = st.sidebar.checkbox("🔘 Enable Stock List Input", value=False)

            if use_input_field:
                raw_input = st.sidebar.text_input("Enter Stock List (comma-separated)", value="")
                if raw_input.strip():
                    input_stocks = [s.strip().upper() for s in raw_input.split(",")]
                    filtered_stocks = [s for s in stocks if s.upper() in input_stocks]
                    if not filtered_stocks:
                        st.sidebar.warning("⚠️ No matching stocks found from your input.")
                    selected_stocks = st.sidebar.multiselect("Select Stocks", filtered_stocks, default=filtered_stocks)
                else:
                    st.sidebar.info("ℹ️ Please input stock symbols to filter the selection.")
                    selected_stocks = st.sidebar.multiselect("Select Stocks", [], default=[])
            else:
                selected_stocks = st.sidebar.multiselect("Select Stocks", stocks, default=[])

            data_columns = ["Open", "High", "Low", "Close", "Volume", "Value", "VWAP"]
            if "selected_columns" not in st.session_state:
                st.session_state.selected_columns = ["Close"]
            selected_columns = st.sidebar.multiselect("Select Data Columns", data_columns, default=st.session_state.selected_columns)
            st.session_state.selected_columns = selected_columns

            available_columns = df.columns.tolist()
            valid_selected_columns = [col for col in selected_columns if col in available_columns]
            missing_columns = [col for col in selected_columns if col not in available_columns]
            if missing_columns:
                st.warning(f"⚠️ The following column(s) are not in the database: {', '.join(missing_columns)}")
                if not valid_selected_columns:
                    st.stop()

            filtered_df = df[
                (df["Date"] >= date_range[0]) &
                (df["Date"] <= date_range[1]) &
                (df["Stock"].isin(selected_stocks))
            ][["Date", "Stock"] + valid_selected_columns].copy()

            if filtered_df.empty:
                st.warning("⚠️ No data to display. Please make sure to select at least one stock using the input or multiselect.")
                st.stop()

            filtered_df.sort_values(["Stock", "Date"], inplace=True)
            filtered_df[valid_selected_columns] = filtered_df.groupby("Stock")[valid_selected_columns].transform(lambda x: x.ffill())

            st.subheader("📑 Filtered Dataset")

            if len(valid_selected_columns) == 1:
                value_col = valid_selected_columns[0]
                pivot_df = filtered_df.pivot(index="Date", columns="Stock", values=value_col).sort_index()
                st.dataframe(pivot_df)

                st.sidebar.markdown("---")
                st.sidebar.subheader("📊 Analysis Options")
                selected_analyses = st.sidebar.multiselect(
                    "Select Analysis / Statistics to Perform",
                    ["Daily Return", "Volatility", "Correlation", "Regression"]
                )
                if selected_analyses:
                    st.subheader("📈 Analysis / Statistics")

                if "Daily Return" in selected_analyses:
                    daily_return = pivot_df.pct_change().dropna()
                    st.markdown("**📈 Daily Return (% Change)**")
                    st.dataframe(daily_return)

                if "Volatility" in selected_analyses:
                    if 'daily_return' not in locals():
                        daily_return = pivot_df.pct_change().dropna()
                    volatility = daily_return.std()
                    st.markdown("**📉 Volatility (Standard Deviation of Daily Return)**")
                    st.dataframe(volatility.to_frame(name="Volatility"))

                if "Correlation" in selected_analyses:
                    if 'daily_return' not in locals():
                        daily_return = pivot_df.pct_change().dropna()
                    correlation = daily_return.corr()
                    st.markdown("**🔗 Correlation Matrix**")
                    st.dataframe(correlation)

                if "Regression" in selected_analyses:
                    if 'daily_return' not in locals():
                        daily_return = pivot_df.pct_change().dropna()
                    st.markdown("**📐 Regression vs. Benchmark**")
                    benchmark_stock = st.selectbox("Select Benchmark Stock", pivot_df.columns)
                    import statsmodels.api as sm
                    regressions = []
                    for stock in daily_return.columns:
                        if stock == benchmark_stock:
                            continue
                        X = daily_return[benchmark_stock]
                        y = daily_return[stock]
                        X = sm.add_constant(X)
                        model = sm.OLS(y, X).fit()
                        regressions.append({
                            "Stock": stock,
                            "Beta": model.params[benchmark_stock],
                            "Alpha": model.params["const"],
                            "R-squared": model.rsquared
                        })
                    reg_df = pd.DataFrame(regressions)
                    st.dataframe(reg_df)

            else:
                reshaped = []
                for col in valid_selected_columns:
                    temp = filtered_df.pivot(index="Date", columns="Stock", values=col)
                    temp.columns = [f"{stock}_{col}" for stock in temp.columns]
                    reshaped.append(temp)
                combined_df = pd.concat(reshaped, axis=1).sort_index()
                st.dataframe(combined_df)
